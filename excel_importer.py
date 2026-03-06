"""Excel race card import for Fantasy Furlong."""
from __future__ import annotations

import re
from datetime import date, datetime
from zoneinfo import ZoneInfo
from typing import Any

from openpyxl import load_workbook


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def parse_odds_to_decimal(raw: str | None) -> float | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if "/" in text:
        try:
            num, den = text.split("/", 1)
            return (float(num) / float(den)) + 1.0
        except Exception:
            return None
    try:
        value = float(text)
        return value if value > 1 else None
    except Exception:
        return None


def parse_race_header(label: str) -> tuple[int, str] | None:
    m = re.search(r"race\s*(\d+)\s*[-–]\s*(\d{1,2}:\d{2})", label, flags=re.IGNORECASE)
    if not m:
        return None
    return int(m.group(1)), m.group(2)


def import_racecard_xlsx(file_path: str, meeting_name: str, course: str, date_local: str, config: dict[str, Any]) -> tuple[dict[str, Any], dict[str, Any], list[str]]:
    wb = load_workbook(filename=file_path, data_only=True)
    ws = wb.active
    tz = ZoneInfo(config["timezone"])

    meeting_id = f"meet_{date_local}_{slugify(course)}"
    snapshot_id = f"odds_{date_local}_0900"

    races: list[dict[str, Any]] = []
    warnings: list[str] = []
    current_race: dict[str, Any] | None = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        horse_name = row[0]
        odds_value = row[1] if len(row) > 1 else None
        if horse_name is None or str(horse_name).strip() == "":
            continue
        name_text = str(horse_name).strip()

        parsed_header = parse_race_header(name_text)
        if parsed_header:
            race_index, off_time = parsed_header
            local_dt = datetime.combine(date.fromisoformat(date_local), datetime.strptime(off_time, "%H:%M").time(), tz)
            off_utc = local_dt.astimezone(ZoneInfo("UTC")).isoformat()
            race_id = f"rac_{meeting_id}_{race_index}"
            current_race = {
                "race_id": race_id,
                "meeting_id": meeting_id,
                "race_index": race_index,
                "name": f"Race {race_index}",
                "off_time_local": off_time,
                "scheduled_off_dt_utc": off_utc,
                "status": "open",
                "locked": False,
                "rescore_count": 0,
                "runners": [],
                "results": None,
                "day": 1,
            }
            races.append(current_race)
            continue

        if current_race is None:
            continue

        runner_no = len(current_race["runners"]) + 1
        race_id = current_race["race_id"]
        dec = parse_odds_to_decimal(odds_value)
        odds_status = "ok" if dec else "missing"
        runner = {
            "runner_id": f"run_{race_id}_{runner_no}",
            "race_id": race_id,
            "number": runner_no,
            "horse_name": name_text,
            "market_odds": str(odds_value).strip() if odds_value is not None else None,
            "market_decimal": dec,
            "fair_decimal": None,
            "place_decimal_fair": None,
            "odds_status": odds_status,
            "scoreable": bool(dec),
            "allow_pick": True,
        }
        current_race["runners"].append(runner)

    for race in races:
        valid = [r for r in race["runners"] if r["market_decimal"]]
        if len(valid) < 2:
            race["status"] = "awaiting_odds"
            warnings.append(f"{race['name']} has fewer than two valid odds entries.")

    meeting = {
        "meeting_id": meeting_id,
        "name": meeting_name,
        "course": course,
        "date_local": date_local,
        "timezone": config["timezone"],
        "races": races,
        "players": [],
        "status": "racecard_loaded" if races else "empty",
        "snapshot_id": snapshot_id,
    }

    snapshot = {
        "snapshot_id": snapshot_id,
        "meeting_id": meeting_id,
        "captured_at_utc": datetime.now(ZoneInfo("UTC")).isoformat(),
        "races": [
            {
                "race_id": race["race_id"],
                "runners": [
                    {
                        "runner_id": r["runner_id"],
                        "market_odds": r["market_odds"],
                        "market_decimal": r["market_decimal"],
                    }
                    for r in race["runners"]
                ],
            }
            for race in races
        ],
    }
    return meeting, snapshot, warnings
