"""Shared backend API for Fantasy Furlong web clients.

Provides cached/batched API pulls and lightweight app endpoints consumed by the PWA.
"""
from __future__ import annotations

import base64
import hashlib
import hmac
import json
import os
import secrets
import threading
import time
from copy import deepcopy
from datetime import date, datetime, timedelta, timezone
from json import JSONDecodeError
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import portalocker
from fastapi import Depends, FastAPI, File, HTTPException, Query, Request, Response, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

from engine import odds_multiplier, paid_places_for_field_size, recompute_fair_odds_for_race, rescore_session
from excel_importer import parse_odds_to_decimal, parse_race_header
from master_config_loader import load_master_config, load_racing_api_credentials
from storage import ensure_dirs, save_session, SESSIONS
from theracingapi_adapter import adapt_racecards_to_session, apply_results_to_session, find_course_id
from numeric_utils import safe_float, safe_int
from theracingapi_client import (
    TheRacingApiClient,
    TheRacingApiRequestError,
    cache_clear,
    cache_stats,
    set_runtime_endpoint_ttls,
    vendor_call_stats,
)

ROOT = Path(__file__).parent
DIST = ROOT / "dist"
DIST_ASSETS = DIST / "assets"
DATA = ROOT / "data"
WEB = DATA / "web"
API_PULLS = DATA / "api_pulls"
ADMIN_AUDIT_LOG = DATA / "admin_audit.log"
ADMIN_STATE_PATH = DATA / "admin_state.json"
INVITES_PATH = DATA / "invites.json"
UPLOAD_PREVIEW_PATH = DATA / "upload_preview.json"
RACEDAYS_CONFIG_PATH = DATA / "racedays.json"
ADMIN_COOKIE = "ff_admin"
USER_COOKIE = "ff_user"

FF_FESTIVAL_MODE = os.getenv("FF_FESTIVAL_MODE", "1") == "1"
FF_FESTIVAL_LENGTH_HOURS = safe_int(os.getenv("FF_FESTIVAL_LENGTH_HOURS", "96")) or 96
FF_COURSES_TTL_SECONDS = safe_int(os.getenv("FF_COURSES_TTL_SECONDS", str(96 * 60 * 60))) or (96 * 60 * 60)
FF_RACECARDS_TTL_SECONDS = safe_int(os.getenv("FF_RACECARDS_TTL_SECONDS", "120")) or 120
FF_RESULTS_TTL_SECONDS = safe_int(os.getenv("FF_RESULTS_TTL_SECONDS", "60")) or 60
FF_RACECARDS_TTL_SECONDS_ACTIVE = safe_int(os.getenv("FF_RACECARDS_TTL_SECONDS_ACTIVE", "60")) or 60
FF_RESULTS_TTL_SECONDS_ACTIVE = safe_int(os.getenv("FF_RESULTS_TTL_SECONDS_ACTIVE", "30")) or 30
FF_ROUTE_CACHE_SECONDS = safe_int(os.getenv("FF_ROUTE_CACHE_SECONDS", "10")) or 10
FF_USER_SESSION_MAX_AGE_SECONDS = max(24 * 3600, FF_FESTIVAL_LENGTH_HOURS * 3600)


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _resolve_api_day_and_local_date(day: str, local_tz) -> tuple[str, str]:
    now_local = _utc_now().astimezone(local_tz).date()
    token = (day or "today").strip().lower()
    if token == "today":
        return "today", now_local.isoformat()
    if token == "tomorrow":
        d = now_local + timedelta(days=1)
        return "tomorrow", d.isoformat()
    try:
        requested = date.fromisoformat(token)
    except ValueError:
        raise ValueError("FF_RACECARD_DAY must be 'today', 'tomorrow', or YYYY-MM-DD")
    if requested == now_local:
        return "today", requested.isoformat()
    if requested == now_local + timedelta(days=1):
        return "tomorrow", requested.isoformat()
    raise ValueError("FF_RACECARD_DAY must map to local today or tomorrow")


def _resolve_timezone(name: str):
    try:
        return ZoneInfo(name)
    except ZoneInfoNotFoundError:
        return timezone.utc


def _default_racedays_payload() -> dict[str, Any]:
    return {
        "raceDays": [
            {"course": "Cheltenham", "date": "2026-03-10"},
            {"course": "Cheltenham", "date": "2026-03-11"},
            {"course": "Cheltenham", "date": "2026-03-12"},
            {"course": "Cheltenham", "date": "2026-03-13"},
        ]
    }


def _parse_configured_racedays(payload: Any) -> list[dict[str, Any]]:
    rows = payload.get("raceDays", []) if isinstance(payload, dict) else []
    parsed: list[dict[str, Any]] = []
    for idx, row in enumerate(rows):
        if not isinstance(row, dict):
            continue
        slot_raw = row.get("slot", idx + 1)
        try:
            slot = safe_int(slot_raw) or (idx + 1)
        except Exception:
            slot = idx + 1
        course = str(row.get("course", "")).strip()
        day = str(row.get("date", "")).strip()
        label = str(row.get("label", "")).strip()
        if not course or not day:
            continue
        try:
            date.fromisoformat(day)
        except ValueError:
            print(f"[WARN] racedays.json slot {slot} has invalid date '{day}', expected YYYY-MM-DD; ignored")
            continue
        parsed.append({"slot": slot, "course": course, "date": day, "label": label})
    parsed.sort(key=lambda x: safe_int(x.get("slot", 9999)) or 9999)
    return parsed


def _meeting_days_from_racedays(race_days: list[dict[str, Any]]) -> list[str]:
    out: list[str] = []
    for row in race_days:
        label = str(row.get("label", "")).strip()
        if label:
            out.append(label)
        else:
            out.append(f"{row.get('course', '')} ({row.get('date', '')})")
    return out


def _annotate_next_check(day_row: dict[str, Any], refresh_seconds: int) -> dict[str, Any]:
    row = dict(day_row)
    next_check = row.get("next_check_utc")
    if not next_check:
        base = row.get("last_refresh")
        try:
            if base:
                next_dt = datetime.fromisoformat(str(base)) + timedelta(seconds=refresh_seconds)
            else:
                next_dt = _utc_now() + timedelta(seconds=refresh_seconds)
            row["next_check_utc"] = next_dt.isoformat()
        except Exception:
            row["next_check_utc"] = (_utc_now() + timedelta(seconds=refresh_seconds)).isoformat()
    return row


def _format_weight_st_lbs(value: Any) -> str:
    if value in (None, ""):
        return ""
    raw = str(value).strip()
    if not raw:
        return ""
    if "-" in raw and "st" not in raw.lower() and "lbs" not in raw.lower():
        left, right = [x.strip() for x in raw.split("-", 1)]
        st = safe_int(left)
        lbs = safe_int(right)
        if st is not None and lbs is not None and 0 <= lbs < 14:
            return f"{st}st {lbs}lbs"
    as_lbs = safe_int(raw)
    if as_lbs is None:
        return ""
    stones = as_lbs // 14
    pounds = as_lbs % 14
    return f"{stones}st {pounds}lbs"


def _read_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except JSONDecodeError:
        bad = path.with_suffix(path.suffix + ".bad")
        try:
            path.rename(bad)
        except Exception:
            pass
        return default


def _atomic_write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lock_path = path.with_suffix(path.suffix + ".lock")
    with portalocker.Lock(lock_path, timeout=5):
        tmp = path.with_suffix(path.suffix + ".tmp")
        tmp.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        tmp.replace(path)


def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def _audit(action: str, detail: dict[str, Any]) -> None:
    ADMIN_AUDIT_LOG.parent.mkdir(parents=True, exist_ok=True)
    row = {
        "at": _utc_now().isoformat(),
        "action": action,
        "detail": detail,
    }
    with ADMIN_AUDIT_LOG.open("a", encoding="utf-8") as f:
        f.write(json.dumps(row) + "\n")


def _b64url(data: bytes) -> str:
    return base64.urlsafe_b64encode(data).decode("ascii").rstrip("=")


def _b64url_decode(text: str) -> bytes:
    pad = "=" * ((4 - len(text) % 4) % 4)
    return base64.urlsafe_b64decode(text + pad)


def _token_sign(secret: str, payload: dict[str, Any]) -> str:
    body = _b64url(json.dumps(payload, separators=(",", ":")).encode("utf-8"))
    sig = hmac.new(secret.encode("utf-8"), body.encode("ascii"), hashlib.sha256).digest()
    return f"{body}.{_b64url(sig)}"


def _token_verify(secret: str, token: str) -> dict[str, Any] | None:
    try:
        body, sig = token.split(".", 1)
        expect = hmac.new(secret.encode("utf-8"), body.encode("ascii"), hashlib.sha256).digest()
        if not hmac.compare_digest(expect, _b64url_decode(sig)):
            return None
        payload = json.loads(_b64url_decode(body).decode("utf-8"))
        if (safe_int(payload.get("exp", 0)) or 0) < int(time.time()):
            return None
        return payload
    except Exception:
        return None


class PickIn(BaseModel):
    user_id: str
    race_id: str
    runner_id: str


class UserIn(BaseModel):
    display_name: str


class RaceLockIn(BaseModel):
    race_id: str
    locked: bool
    confirm: bool | None = None


class PasswordIn(BaseModel):
    password: str


class ManualResultIn(BaseModel):
    first: str
    second: str
    third: str
    fourth: str
    dnf_ids: list[str] = []
    nr_ids: list[str] = []


class InviteCreateIn(BaseModel):
    count: int
    expires_hours: int


class InviteRevokeIn(BaseModel):
    token: str


class ForceSettleIn(BaseModel):
    race_id: str


class JoinIn(BaseModel):
    token: str
    display_name: str


class AppState:
    def __init__(self) -> None:
        ensure_dirs()
        WEB.mkdir(parents=True, exist_ok=True)
        API_PULLS.mkdir(parents=True, exist_ok=True)
        self.config = load_master_config(ROOT / "Master_config_V1.txt")
        self.tz = self.config.get("timezone", "Europe/London")
        self.local_tz = _resolve_timezone(self.tz)
        self.user, self.password, self.cred_source = load_racing_api_credentials(ROOT, self.config)
        self.target_course = os.getenv("FF_TARGET_COURSE", "Exeter")
        self.refresh_seconds = safe_int(os.getenv("FF_REFRESH_SECONDS", "60")) or 60
        if FF_FESTIVAL_MODE and self.refresh_seconds < 30:
            print("[WARN] FF_REFRESH_SECONDS too low for festival mode; clamped to 30")
            self.refresh_seconds = 30
        self.route_cache_seconds = FF_ROUTE_CACHE_SECONDS
        self.racedays_path = RACEDAYS_CONFIG_PATH
        self.configured_race_days = self._load_configured_race_days()
        self.race_day_states: list[dict[str, Any]] = []
        self.course_id: str | int | None = None
        self._lock = threading.Lock()
        self._cache: dict[str, Any] = {
            "session": None,
            "meeting": None,
            "races": [],
            "last_refresh": None,
            "last_error": None,
            "generated_at_utc": None,
            "last_successful_vendor_pull_at": None,
            "stale": False,
            "source_age_seconds": 0,
        }
        self._route_cache: dict[str, dict[str, Any]] = {}
        self._route_cache_hits = 0
        self._route_cache_misses = 0
        self._stop = threading.Event()
        self._thread: threading.Thread | None = None
        self._refresh_lock = threading.Lock()
        self.admin_password = os.getenv("FF_ADMIN_PASSWORD", "")
        self.admin_password_hash = os.getenv("FF_ADMIN_PASSWORD_HASH", "")
        self.admin_secret = os.getenv("FF_ADMIN_SECRET") or secrets.token_urlsafe(32)
        if not os.getenv("FF_ADMIN_SECRET"):
            print("[WARN] FF_ADMIN_SECRET missing. Generated ephemeral secret for this process only.")
        set_runtime_endpoint_ttls(racecards_ttl=FF_RACECARDS_TTL_SECONDS, results_ttl=FF_RESULTS_TTL_SECONDS)

    @property
    def users_path(self) -> Path:
        return WEB / "users.json"

    @property
    def picks_path(self) -> Path:
        return WEB / "picks.json"

    def _load_admin_state(self) -> dict[str, Any]:
        default = {"race_locks": {}, "manual_results": {}, "settlement_overrides": {}, "manual_upload": None}
        data = _read_json(ADMIN_STATE_PATH, default)
        if not isinstance(data, dict):
            return default
        for k, v in default.items():
            data.setdefault(k, v)
        return data

    def _save_admin_state(self, payload: dict[str, Any]) -> None:
        _atomic_write_json(ADMIN_STATE_PATH, payload)

    def _route_cache_get(self, key: str) -> Any | None:
        now = time.time()
        row = self._route_cache.get(key)
        if not row or row.get("expires_at", 0) <= now:
            self._route_cache_misses += 1
            return None
        self._route_cache_hits += 1
        return deepcopy(row.get("payload"))

    def _route_cache_set(self, key: str, payload: Any) -> None:
        self._route_cache[key] = {"payload": deepcopy(payload), "expires_at": time.time() + self.route_cache_seconds}

    def _route_cache_invalidate(self) -> None:
        self._route_cache.clear()

    def _apply_dynamic_cache_ttls(self, session: dict[str, Any]) -> None:
        races = session.get("meeting", {}).get("races", [])
        offs: list[datetime] = []
        for race in races:
            dt_text = race.get("scheduled_off_dt_utc")
            if isinstance(dt_text, str):
                try:
                    offs.append(datetime.fromisoformat(dt_text))
                except ValueError:
                    pass
        if not offs:
            set_runtime_endpoint_ttls(racecards_ttl=FF_RACECARDS_TTL_SECONDS, results_ttl=FF_RESULTS_TTL_SECONDS)
            return
        now = _utc_now()
        window_start = min(offs) - timedelta(minutes=30)
        window_end = max(offs) + timedelta(minutes=30)
        active = window_start <= now <= window_end
        if active:
            set_runtime_endpoint_ttls(racecards_ttl=FF_RACECARDS_TTL_SECONDS_ACTIVE, results_ttl=FF_RESULTS_TTL_SECONDS_ACTIVE)
        else:
            set_runtime_endpoint_ttls(racecards_ttl=FF_RACECARDS_TTL_SECONDS, results_ttl=FF_RESULTS_TTL_SECONDS)

    def race_lock_flags(self) -> dict[str, bool]:
        state = self._load_admin_state()
        return {str(k): bool(v) for k, v in state.get("race_locks", {}).items()}

    def set_race_lock_flag(self, race_id: str, locked: bool) -> None:
        with self._lock:
            state = self._load_admin_state()
            state.setdefault("race_locks", {})[str(race_id)] = bool(locked)
            self._save_admin_state(state)

    def _strip_legacy_npc_users(self, users: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
        cleaned: list[dict[str, Any]] = []
        removed = 0
        for row in users:
            uid = str(row.get("id", "")).strip()
            name = str(row.get("displayName", "")).strip()
            avatar = str(row.get("avatar", "")).strip()
            is_admin = bool(row.get("isAdmin", False))
            # Remove legacy scaffold NPCs that were previously auto-seeded as
            # `u_you` + `u_p1..u_p14` / `Player 1..14`.
            is_legacy_seed_id = uid == "u_you" or (uid.startswith("u_p") and uid[3:].isdigit())
            is_legacy_seed_name = name.lower() == "you" or (name.lower().startswith("player ") and name[7:].strip().isdigit())
            if is_legacy_seed_id and (is_legacy_seed_name or avatar in {"P", "Y"} or (uid == "u_you" and is_admin)):
                removed += 1
                continue
            cleaned.append(row)
        return cleaned, removed

    def users(self) -> list[dict[str, Any]]:
        users = _read_json(self.users_path, [])
        if not isinstance(users, list):
            _write_json(self.users_path, [])
            return []
        cleaned, removed = self._strip_legacy_npc_users(users)
        if removed:
            _write_json(self.users_path, cleaned)
        return cleaned

    def picks(self) -> list[dict[str, Any]]:
        return _read_json(self.picks_path, [])


    def _load_configured_race_days(self) -> list[dict[str, Any]]:
        default_payload = _default_racedays_payload()
        if not self.racedays_path.exists():
            _write_json(self.racedays_path, default_payload)
        payload = _read_json(self.racedays_path, default_payload)
        if not isinstance(payload, dict):
            payload = default_payload
            _write_json(self.racedays_path, payload)
        rows = payload.get("raceDays")
        if not isinstance(rows, list):
            payload["raceDays"] = default_payload["raceDays"]
            _write_json(self.racedays_path, payload)
            rows = payload["raceDays"]
        normalized_rows: list[dict[str, Any]] = []
        for row in rows[:7]:
            if isinstance(row, dict):
                normalized_rows.append({
                    "course": str(row.get("course", "")),
                    "date": str(row.get("date", "")),
                })
            else:
                normalized_rows.append({"course": "", "date": ""})
        if len(normalized_rows) < 7:
            normalized_rows.extend({"course": "", "date": ""} for _ in range(len(normalized_rows), 7))
        payload["raceDays"] = normalized_rows
        _write_json(self.racedays_path, payload)
        return _parse_configured_racedays(payload)

    def set_pick(self, pick: PickIn) -> None:
        snap = self.snapshot()
        races = snap.get("session", {}).get("meeting", {}).get("races", [])
        race = next((r for r in races if str(r.get("race_id")) == str(pick.race_id)), None)
        if not race:
            raise HTTPException(status_code=404, detail="Race not found")
        scheduled_off_utc = race.get("scheduled_off_dt_utc")
        if isinstance(scheduled_off_utc, str):
            try:
                if _utc_now() >= datetime.fromisoformat(scheduled_off_utc):
                    raise HTTPException(status_code=409, detail="Error: Race Started, Picks are Locked.")
            except ValueError:
                pass
        items = [p for p in self.picks() if not (p["userId"] == pick.user_id and p["raceId"] == pick.race_id)]
        items.append({"userId": pick.user_id, "raceId": pick.race_id, "runnerId": pick.runner_id})
        _write_json(self.picks_path, items)

    def get_user_by_id(self, user_id: str) -> dict[str, Any] | None:
        return next((u for u in self.users() if str(u.get("id")) == str(user_id)), None)

    def add_user(self, payload: UserIn) -> dict[str, Any]:
        name = payload.display_name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="display_name required")
        users = self.users()
        existing = next((u for u in users if str(u.get("displayName", "")).strip().lower() == name.lower()), None)
        if existing:
            return existing
        uid = f"u_{int(time.time() * 1000)}"
        row = {"id": uid, "displayName": name, "isAdmin": False, "avatar": name[0].upper()}
        users.append(row)
        _write_json(self.users_path, users)
        return row

    def _fetch_racecards_session(self, local_date: str, course: str | None = None) -> tuple[dict[str, Any], str | int]:
        target_course = (course or self.target_course).strip() or self.target_course
        client = TheRacingApiClient(self.user or "", self.password or "")
        if self.course_id is None or target_course != self.target_course:
            courses = client.list_courses(region_codes=["gb"])
            course_id = find_course_id(courses, target_course)
            if target_course == self.target_course:
                self.course_id = course_id
        else:
            course_id = self.course_id
        now_local = _utc_now().astimezone(self.local_tz).date()
        req = date.fromisoformat(local_date)
        if req == now_local:
            day_token = "today"
        elif req == now_local + timedelta(days=1):
            day_token = "tomorrow"
        else:
            # Allow direct ISO date lookups so future configured festival days can
            # load as soon as the vendor publishes racecards.
            day_token = local_date
        racecards = client.fetch_racecards_standard(day=day_token, course_ids=[course_id], region_codes=["gb"])
        session = adapt_racecards_to_session(
            racecards_json=racecards,
            target_course_id=course_id,
            target_course_name=target_course,
            target_date_local=local_date,
            timezone=self.tz,
            config=self.config,
        )
        return session, course_id



    def _load_persisted_session_candidates(self) -> list[dict[str, Any]]:
        sessions: list[dict[str, Any]] = []
        if not SESSIONS.exists():
            return sessions
        files = sorted(SESSIONS.glob('*.json'), key=lambda f: f.stat().st_mtime, reverse=True)
        for fp in files:
            try:
                payload = json.loads(fp.read_text(encoding='utf-8'))
            except Exception:
                continue
            if isinstance(payload, dict) and isinstance(payload.get('meeting'), dict):
                sessions.append(payload)
        return sessions

    def refresh_once(self, force: bool, mode: str = "full") -> dict[str, Any]:
        if not (self.user and self.password):
            with self._lock:
                self._cache["last_error"] = "Missing API credentials"
                if self._cache.get("meeting") is not None:
                    self._cache["stale"] = True
                    last_ok = self._cache.get("last_successful_vendor_pull_at")
                    if last_ok:
                        self._cache["source_age_seconds"] = int((_utc_now() - datetime.fromisoformat(last_ok)).total_seconds())
            raise HTTPException(status_code=503, detail="Missing API credentials")
        acquired = self._refresh_lock.acquire(blocking=False)
        if not acquired:
            raise HTTPException(status_code=409, detail="Refresh already in progress")
        started = _utc_now()
        errors: list[str] = []
        fetched = {"racecards": 0, "results": 0}
        try:
            self.configured_race_days = self._load_configured_race_days()
            now_iso = _utc_now().isoformat()
            day_states: list[dict[str, Any]] = []
            sessions: list[dict[str, Any]] = []
            all_results: list[dict[str, Any]] = []
            today_local = _utc_now().astimezone(self.local_tz).date()
            persisted_candidates = self._load_persisted_session_candidates()
            persisted_races: list[dict[str, Any]] = []
            for prev in persisted_candidates:
                for race in prev.get("meeting", {}).get("races", []):
                    if isinstance(race, dict):
                        persisted_races.append(deepcopy(race))

            def _persisted_for_day(slot: int, course: str, local_day: str) -> list[dict[str, Any]]:
                out: list[dict[str, Any]] = []
                seen: set[str] = set()
                for race in persisted_races:
                    if (safe_int(race.get("_ff_slot")) or -1) != slot:
                        continue
                    if str(race.get("_ff_course", "")).strip().lower() != course.strip().lower():
                        continue
                    if str(race.get("_ff_date", "")).strip() != local_day:
                        continue
                    rid = str(race.get("race_id", "")).strip()
                    if rid and rid in seen:
                        continue
                    if rid:
                        seen.add(rid)
                    out.append(deepcopy(race))
                return out

            for row in self.configured_race_days:
                slot = safe_int(row.get("slot", len(day_states) + 1)) or (len(day_states) + 1)
                local_day = str(row.get("date", ""))
                course = str(row.get("course", "")).strip()
                label = str(row.get("label", "")).strip()
                state = {
                    "slot": slot,
                    "course": course,
                    "date": local_day,
                    "label": label,
                    "status": "pending",
                    "races": [],
                    "last_refresh": None,
                    "last_error": None,
                    "next_check_utc": None,
                }
                try:
                    req_day = date.fromisoformat(local_day)
                except ValueError:
                    state["status"] = "error"
                    state["last_error"] = f"Invalid configured date: {local_day}"
                    day_states.append(state)
                    continue

                existing_day_races = _persisted_for_day(slot, course, local_day)

                try:
                    session, course_id = self._fetch_racecards_session(local_day, course)
                    for race in session.get("meeting", {}).get("races", []):
                        race["_ff_day_index"] = slot - 1
                        race["_ff_course"] = course
                        race["_ff_date"] = local_day
                        race["_ff_slot"] = slot
                    sessions.append(session)
                    fetched["racecards"] += len(session.get("meeting", {}).get("races", []))

                    day_results: list[dict[str, Any]] = []
                    if local_day <= today_local.isoformat():
                        try:
                            client = TheRacingApiClient(self.user, self.password)
                            day_results = client.fetch_results(start_date=local_day, end_date=local_day, course=[course_id])
                            fetched["results"] += len(day_results)
                        except TheRacingApiRequestError as exc:
                            if exc.status_code != 422:
                                raise
                    apply_results_to_session(session, day_results, self.config)
                    all_results.extend(day_results)
                    state["status"] = "loaded"
                    state["last_refresh"] = now_iso
                    state["next_check_utc"] = (_utc_now() + timedelta(seconds=self.refresh_seconds)).isoformat()
                    state["races"] = [
                        {
                            "id": str(r.get("race_id", "")),
                            "off_time": r.get("off_time_local", ""),
                            "name": r.get("name", ""),
                            "status": str(r.get("status", "open")),
                        }
                        for r in session.get("meeting", {}).get("races", [])
                    ]
                except Exception as exc:
                    if existing_day_races:
                        state["status"] = "loaded"
                        state["last_error"] = f"{exc} (showing last stored data)"
                        state["races"] = [
                            {
                                "id": str(r.get("race_id", "")),
                                "off_time": r.get("off_time_local", ""),
                                "name": r.get("name", ""),
                                "status": str(r.get("status", "open")),
                            }
                            for r in existing_day_races
                        ]
                    else:
                        state["status"] = "error"
                        state["last_error"] = str(exc)
                    state["next_check_utc"] = (_utc_now() + timedelta(seconds=self.refresh_seconds)).isoformat()
                    errors.append(f"slot {slot}: {exc}")
                day_states.append(state)

            self.race_day_states = [
                _annotate_next_check(d, self.refresh_seconds)
                for d in sorted(day_states, key=lambda d: safe_int(d.get("slot", 9999)) or 9999)
            ]

            combined = deepcopy(persisted_candidates[0]) if persisted_candidates else {"meeting": {"races": []}}
            combined_meeting = combined.setdefault("meeting", {})
            if not combined_meeting.get("meeting_id"):
                combined_meeting["meeting_id"] = "festival_current"
            combined_meeting.setdefault("course", self.target_course)
            combined_meeting.setdefault("source", "api")
            merged_races: dict[str, dict[str, Any]] = {}
            for race in persisted_races:
                rid = str(race.get("race_id", "")).strip()
                if rid:
                    merged_races[rid] = deepcopy(race)
            for ses in sessions:
                for race in ses.get("meeting", {}).get("races", []):
                    rid = str(race.get("race_id", "")).strip()
                    if rid:
                        merged_races[rid] = deepcopy(race)
            combined_meeting["races"] = sorted(merged_races.values(), key=lambda r: r.get("scheduled_off_dt_utc", ""))

            if combined_meeting.get("races"):
                self._apply_dynamic_cache_ttls(combined)
                save_session(combined)
            if sessions:
                stamp = _utc_now().strftime("%Y%m%dT%H%M%SZ")
                _write_json(API_PULLS / f"racecards_{stamp}.json", combined)
                _write_json(API_PULLS / f"results_{stamp}.json", all_results)

            meeting, races = self._to_frontend_model(combined)
            with self._lock:
                self._cache.update({
                    "session": combined,
                    "meeting": meeting,
                    "races": races,
                    "last_refresh": now_iso,
                    "last_error": "; ".join(errors) if errors else None,
                    "generated_at_utc": now_iso,
                    "last_successful_vendor_pull_at": now_iso if sessions else self._cache.get("last_successful_vendor_pull_at"),
                    "stale": False,
                    "source_age_seconds": 0,
                })
            self._route_cache_invalidate()
            return {
                "ok": True,
                "mode": mode,
                "started_at": started.isoformat(),
                "finished_at": _utc_now().isoformat(),
                "fetched_counts": fetched,
                "errors": errors,
            }
        except HTTPException:
            raise
        except Exception as exc:
            with self._lock:
                self._cache["last_error"] = str(exc)
                if self._cache.get("meeting") is not None:
                    self._cache["stale"] = True
                    last_ok = self._cache.get("last_successful_vendor_pull_at")
                    if last_ok:
                        self._cache["source_age_seconds"] = int((_utc_now() - datetime.fromisoformat(last_ok)).total_seconds())
            if force and self._cache.get("meeting") is None:
                raise HTTPException(status_code=500, detail=f"refresh failed: {exc}")
            return {"ok": False, "mode": mode, "started_at": started.isoformat(), "finished_at": _utc_now().isoformat(), "fetched_counts": fetched, "errors": [str(exc)]}
        finally:
            self._refresh_lock.release()

    def start(self) -> None:
        # Avoid blocking app startup on vendor I/O. Railway healthchecks can
        # fail if startup waits for upstream API calls/credentials before the
        # server begins accepting requests.
        self._thread = threading.Thread(target=self._loop, daemon=True)
        self._thread.start()

    def stop(self) -> None:
        self._stop.set()

    def _loop(self) -> None:
        while not self._stop.is_set():
            try:
                self.refresh_once(force=False)
            except HTTPException:
                pass
            self._stop.wait(self.refresh_seconds)

    def snapshot(self) -> dict[str, Any]:
        with self._lock:
            return deepcopy(self._cache)

    def _to_frontend_model(self, session: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        meeting = session.get("meeting", {})
        races = sorted(meeting.get("races", []), key=lambda r: r.get("scheduled_off_dt_utc", ""))
        local_day = _utc_now().astimezone(self.local_tz).date().isoformat()
        configured_days = self.race_day_states or [
            {
                "slot": safe_int(d.get("slot", i + 1)) or (i + 1),
                "course": d.get("course", ""),
                "date": d.get("date", ""),
                "label": d.get("label", ""),
                "status": "pending",
                "races": [],
                "last_refresh": None,
                "last_error": None,
                "next_check_utc": (_utc_now() + timedelta(seconds=self.refresh_seconds)).isoformat(),
            }
            for i, d in enumerate(self.configured_race_days)
        ]
        configured_days = [_annotate_next_check(d, self.refresh_seconds) for d in configured_days]
        out_meeting = {
            "id": meeting.get("meeting_id", "meeting"),
            "course": meeting.get("course", self.target_course),
            "festival": "Fantasy Furlong",
            "days": _meeting_days_from_racedays(configured_days),
            "raceDays": configured_days,
            "snapshotLocked": False,
            "refreshIntervalSeconds": self.refresh_seconds,
        }
        out_races: list[dict[str, Any]] = []
        persisted_locks = self.race_lock_flags()
        base_win = self.config["scoring"]["base_points"]["win"]
        base_place = self.config["scoring"]["base_points"]["place"]
        admin_state = self._load_admin_state()
        manual_results = admin_state.get("manual_results", {})
        for idx, race in enumerate(races, start=1):
            day_slot = safe_int(race.get("_ff_slot"))
            expected_day = None
            if day_slot is not None:
                expected_day = next((d for d in configured_days if (safe_int(d.get("slot")) or -1) == day_slot), None)
            race_course = str(race.get("_ff_course", "")).strip().lower()
            race_date = str(race.get("_ff_date", "")).strip()
            if expected_day:
                exp_course = str(expected_day.get("course", "")).strip().lower()
                exp_date = str(expected_day.get("date", "")).strip()
                if race_course != exp_course or race_date != exp_date:
                    continue
            recompute_fair_odds_for_race(race, self.config)
            result = race.get("results") or {}
            override = manual_results.get(str(race.get("race_id")))
            if override:
                placements = [override.get("first"), override.get("second"), override.get("third"), override.get("fourth")]
                dnf_ids = {str(x) for x in override.get("dnf_ids", [])}
            else:
                placements = [str(x) for x in result.get("placements", []) if x]
                dnf_ids = {str(x) for x in result.get("dnf_runner_ids", [])}
            paid_places = paid_places_for_field_size(len(race.get("runners", [])), self.config)
            frunners = []
            for num, rn in enumerate(race.get("runners", []), start=1):
                rid = str(rn.get("runner_id"))
                pos = placements.index(rid) + 1 if rid in placements else None
                fair = rn.get("fair_decimal")
                place_fair = rn.get("place_decimal_fair")
                points_win = round(base_win * odds_multiplier(fair, "sqrt"), 2) if fair else 0
                points_place = round(base_place * odds_multiplier(place_fair, "sqrt"), 2) if place_fair else 0
                points_awarded = points_win if pos == 1 else (points_place if pos and pos <= paid_places else (self.config["scoring"]["dnf_penalty"]["points"] if rid in dnf_ids else 0))
                age = rn.get("age")
                sex = str(rn.get("sex_code", "")).strip()
                details = " ".join([x for x in [f"{age}yo" if age not in (None, "") else "", sex] if x]).strip()
                weight_text = _format_weight_st_lbs(rn.get("weight") or rn.get("lbs"))
                sire = str(rn.get("sire", "")).strip()
                dam = str(rn.get("dam", "")).strip()
                breeding = " - ".join([x for x in [sire, dam] if x])
                quotes = rn.get("quotes") if isinstance(rn.get("quotes"), list) else []
                frunners.append({
                    "id": rid,
                    "raceId": str(race.get("race_id")),
                    "horseName": rn.get("horse_name", f"Runner {num}"),
                    "draw": num,
                    "jockey": str(rn.get("jockey", "")),
                    "trainer": str(rn.get("trainer", "")),
                    "owner": str(rn.get("owner", "")),
                    "weight": weight_text,
                    "officialRating": safe_int(rn.get("ofr")) or 0,
                    "form": str(rn.get("form", "")),
                    "details": details,
                    "breeding": breeding,
                    "sire": sire,
                    "dam": dam,
                    "rawOdds": rn.get("market_decimal") or 0,
                    "fairOdds": fair or 0,
                    "isNR": False,
                    "finishPosition": pos,
                    "finishStatus": "dnf" if rid in dnf_ids else "finished",
                    "lengths": None,
                    "pointsWin": round(points_win, 2),
                    "pointsPlace": round(points_place, 2),
                    "pointsAwarded": round(points_awarded, 2),
                    "silkColors": ["#14532d", "#facc15"],
                    "silkUrl": rn.get("silk_url"),
                    "quotes": [str(q) for q in quotes if q],
                    "quote": str(quotes[0]) if quotes else None,
                })
            rstatus = str(race.get("status", "open"))
            status_map = {"open": "declared", "locked": "off", "settling": "off", "awaiting_odds": "declared", "settled": "result"}
            scheduled_off_utc = race.get("scheduled_off_dt_utc")
            off_elapsed = False
            if isinstance(scheduled_off_utc, str):
                try:
                    off_elapsed = _utc_now() >= datetime.fromisoformat(scheduled_off_utc)
                except ValueError:
                    pass
            locked_flag = off_elapsed or bool(persisted_locks.get(str(race.get("race_id")), False))
            day_index = (safe_int(race.get("_ff_day_index")) if race.get("_ff_day_index") not in (None, "", "-") else None)
            if day_index is None:
                day_index = (safe_int(race.get("_ff_slot")) or 1) - 1
            out_races.append({"id": str(race.get("race_id")), "meetingId": out_meeting["id"], "dayIndex": day_index, "raceName": race.get("name", f"Race {idx}"), "offTime": race.get("off_time_local", "00:00"), "distanceMiles": "", "fieldSize": len(frunners), "status": status_map.get(rstatus, "declared"), "settled": rstatus == "settled", "raceNumber": idx, "runners": frunners, "locked": locked_flag})
        manual_upload = admin_state.get("manual_upload")
        if manual_upload and not out_races:
            out_meeting["course"] = manual_upload.get("course", out_meeting["course"])
            for idx, race in enumerate(manual_upload.get("races", []), start=1):
                runners = []
                for i, rn in enumerate(race.get("runners", []), start=1):
                    runners.append({"id": f"{race['race_id']}_r{i}", "raceId": race["race_id"], "horseName": rn.get("horse_name", f"Runner {i}"), "draw": i, "jockey": "", "trainer": "", "owner": "", "weight": "", "officialRating": 0, "form": "", "details": "", "breeding": "", "rawOdds": rn.get("decimal", 0), "fairOdds": rn.get("decimal", 0), "isNR": False, "pointsWin": 0, "pointsPlace": 0, "silkColors": ["#14532d", "#facc15"]})
                out_races.append({"id": race["race_id"], "meetingId": out_meeting["id"], "dayIndex": 0, "raceName": race.get("name", f"Race {idx}"), "offTime": race.get("off_time", "00:00"), "distanceMiles": "", "fieldSize": len(runners), "status": "declared", "settled": False, "raceNumber": idx, "runners": runners, "locked": False})
        return out_meeting, out_races

    def leaderboard(self) -> list[dict[str, Any]]:
        snap = self.snapshot()
        session = snap.get("session")
        if not session:
            return []
        users = self.users()
        picks = self.picks()
        temp = deepcopy(session)
        temp["meeting"]["players"] = [{"player_id": u["id"], "name": u["displayName"]} for u in users]
        allowed_players = {u["id"] for u in users}
        allowed_races = {str(r.get("race_id")) for r in temp.get("meeting", {}).get("races", [])}
        temp["picks"] = [
            {
                "pick_id": f"pick_{p['userId']}_{p['raceId']}",
                "meeting_id": temp["meeting"]["meeting_id"],
                "race_id": p["raceId"],
                "runner_id": p["runnerId"],
                "player_id": p["userId"],
                "locked": False,
            }
            for p in picks
            if p.get("userId") in allowed_players and str(p.get("raceId")) in allowed_races
        ]
        rescore_session(temp, self.config)
        by_user = {row["player_id"]: row for row in temp.get("leaderboard", [])}
        ordered = sorted(users, key=lambda u: -(by_user.get(u["id"], {}).get("points", 0)))
        return [{"userId": u["id"], "displayName": u["displayName"], "totalPoints": round(by_user.get(u["id"], {}).get("points", 0), 2), "dayPoints": [round(by_user.get(u["id"], {}).get("points", 0), 2)], "wins": safe_int(by_user.get(u["id"], {}).get("wins", 0)) or 0, "places": safe_int(by_user.get(u["id"], {}).get("placings", 0)) or 0, "dnfCount": safe_int(by_user.get(u["id"], {}).get("dnfs", 0)) or 0, "position": idx + 1} for idx, u in enumerate(ordered)]


STATE = AppState()
app = FastAPI(title="Fantasy Furlong Backend", version="1.0.0")


def _spa_enabled() -> bool:
    return DIST.exists() and (DIST / "index.html").exists()


def _serve_spa_path(path: str = ""):
    if not _spa_enabled():
        return JSONResponse({"detail": "Frontend not built. Run npm run build."}, status_code=404)
    rel = path.strip("/")
    target = DIST / rel if rel else DIST / "index.html"
    if target.exists() and target.is_file():
        return FileResponse(target)
    return FileResponse(DIST / "index.html")


if DIST_ASSETS.exists():
    app.mount("/assets", StaticFiles(directory=str(DIST_ASSETS)), name="assets")



def _is_admin_password_valid(password: str) -> bool:
    if STATE.admin_password_hash:
        digest = hashlib.sha256(password.encode("utf-8")).hexdigest()
        return hmac.compare_digest(digest, STATE.admin_password_hash)
    if STATE.admin_password:
        return hmac.compare_digest(password, STATE.admin_password)
    return False


def require_admin(request: Request) -> dict[str, Any]:
    token = request.cookies.get(ADMIN_COOKIE)
    if not token:
        raise HTTPException(status_code=401, detail="Admin login required")
    payload = _token_verify(STATE.admin_secret, token)
    if not payload:
        raise HTTPException(status_code=401, detail="Invalid admin session")
    return payload


def require_user(request: Request) -> dict[str, Any]:
    user_id = request.cookies.get(USER_COOKIE)
    if not user_id:
        raise HTTPException(status_code=401, detail="Not authenticated")
    user = STATE.get_user_by_id(user_id)
    if not user:
        raise HTTPException(status_code=401, detail="Unknown user")
    return user


@app.on_event("startup")
def _startup() -> None:
    STATE.start()


@app.on_event("shutdown")
def _shutdown() -> None:
    STATE.stop()


@app.get("/health")
def health() -> dict[str, Any]:
    snap = STATE.snapshot()
    stats = cache_stats()
    return {
        "ok": snap.get("meeting") is not None,
        "last_refresh": snap.get("last_refresh"),
        "last_error": snap.get("last_error"),
        "credentials_source": STATE.cred_source,
        "cache_stats": {
            "hits": stats.get("hits", 0),
            "misses": stats.get("misses", 0),
            "size": stats.get("size", 0),
            "ttl_by_endpoint": stats.get("ttl_by_endpoint", {}),
            "route_cache_hits": STATE._route_cache_hits,
            "route_cache_misses": STATE._route_cache_misses,
        },
        "vendor_call_stats": vendor_call_stats(),
        "snapshot_meta": {
            "generated_at_utc": snap.get("generated_at_utc"),
            "stale": bool(snap.get("stale", False)),
            "source_age_seconds": safe_int(snap.get("source_age_seconds", 0)) or 0,
            "last_successful_vendor_pull_at": snap.get("last_successful_vendor_pull_at"),
        },
    }


@app.post("/api/admin/login")
def admin_login(payload: PasswordIn, response: Response) -> dict[str, Any]:
    if not _is_admin_password_valid(payload.password):
        raise HTTPException(status_code=401, detail="Invalid password")
    now = int(time.time())
    token = _token_sign(STATE.admin_secret, {"sub": "admin", "iat": now, "exp": now + (12 * 3600)})
    response.set_cookie(ADMIN_COOKIE, token, httponly=True, samesite="lax", max_age=12 * 3600, path="/")
    _audit("admin_login", {})
    return {"ok": True}


@app.post("/api/admin/logout")
def admin_logout(response: Response) -> dict[str, Any]:
    response.delete_cookie(ADMIN_COOKIE, path="/")
    _audit("admin_logout", {})
    return {"ok": True}


@app.get("/api/admin/session")
def admin_session(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    return {"ok": True}


@app.post("/api/admin/cache/clear")
def admin_clear_cache(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    cleared = cache_clear()
    STATE._route_cache_invalidate()
    _audit("admin_cache_clear", {"cleared": cleared})
    return {"ok": True, "cleared": cleared}


@app.post("/api/admin/refresh")
def force_refresh(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    result = STATE.refresh_once(force=True, mode="full")
    _audit("admin_refresh_full", result)
    return result


@app.post("/api/admin/refresh/racecards")
def refresh_racecards(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    result = STATE.refresh_once(force=True, mode="racecards")
    _audit("admin_refresh_racecards", result)
    return result


@app.post("/api/admin/refresh/odds")
def refresh_odds(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    result = STATE.refresh_once(force=True, mode="odds")
    _audit("admin_refresh_odds", result)
    return result


@app.post("/api/admin/refresh/results")
def refresh_results(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    result = STATE.refresh_once(force=True, mode="results")
    _audit("admin_refresh_results", result)
    return result


@app.post("/api/admin/refresh/full")
def refresh_full(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    result = STATE.refresh_once(force=True, mode="full")
    _audit("admin_refresh_full", result)
    return result


@app.get("/api/admin/races")
def admin_races(date: str = Query(..., pattern=r"^\d{4}-\d{2}-\d{2}$"), _: dict[str, Any] = Depends(require_admin)) -> list[dict[str, Any]]:
    session, _ = STATE._fetch_racecards_session(date)
    races = session.get("meeting", {}).get("races", [])
    state = STATE._load_admin_state()
    locks = state.get("race_locks", {})
    settled_map = state.get("settlement_overrides", {})
    out = []
    for race in races:
        status = str(race.get("status", "open"))
        completed = status == "settled" or bool((race.get("results") or {}).get("is_official"))
        rid = str(race.get("race_id"))
        off_dt = race.get("scheduled_off_dt_utc", "")
        out.append({"race_id": rid, "course": session.get("meeting", {}).get("course", STATE.target_course), "date": date, "off_dt": off_dt, "off_time": race.get("off_time_local", ""), "race_name": race.get("name", ""), "field_size": len(race.get("runners", [])), "race_status": "result" if completed else status, "locked": bool(locks.get(rid, False)), "settled": bool(settled_map.get(rid, {}).get("official_confirmed", completed))})
    _audit("admin_races_list", {"date": date, "count": len(out)})
    return out


@app.post("/api/admin/race/lock")
def admin_race_lock(payload: RaceLockIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    snap = STATE.snapshot()
    races = snap.get("races", [])
    row = next((r for r in races if str(r.get("id")) == str(payload.race_id)), None)
    race_status = str(row.get("status")) if row else "unknown"
    completed = race_status == "result"
    if payload.locked and not completed:
        raise HTTPException(status_code=409, detail="Race not complete")
    if not payload.locked and not payload.confirm:
        raise HTTPException(status_code=400, detail="Unlock requires confirm=true")
    STATE.set_race_lock_flag(payload.race_id, payload.locked)
    _audit("admin_race_lock", {"race_id": payload.race_id, "locked": payload.locked})
    return {"ok": True, "race_id": payload.race_id, "locked": payload.locked}


@app.post("/api/admin/race/force_settle")
def admin_race_force_settle(payload: ForceSettleIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    race_id = payload.race_id
    state = STATE._load_admin_state()
    so = state.setdefault("settlement_overrides", {})
    row = so.setdefault(race_id, {"settled_at": None, "official_confirmed": False, "rescore_count": 0})
    row["settled_at"] = _utc_now().isoformat()
    row["official_confirmed"] = True
    row["rescore_count"] = (safe_int(row.get("rescore_count", 0)) or 0) + 1
    STATE._save_admin_state(state)
    _audit("admin_force_settle", {"race_id": race_id})
    return {"ok": True, "race_id": race_id, "settled": True}


@app.get("/api/admin/race/{race_id}/runners")
def admin_runners(race_id: str, _: dict[str, Any] = Depends(require_admin)) -> list[dict[str, Any]]:
    snap = STATE.snapshot()
    for race in snap.get("races", []):
        if str(race.get("id")) == str(race_id):
            return [{"horse_id": str(r.get("id")), "horse_name": r.get("horseName", "")} for r in race.get("runners", [])]
    raise HTTPException(status_code=404, detail="Race not found")


@app.post("/api/admin/race/{race_id}/results/provisional")
def admin_provisional_results(race_id: str, payload: ManualResultIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    placings = [payload.first, payload.second, payload.third, payload.fourth]
    if len(set(placings)) != 4:
        raise HTTPException(status_code=400, detail="Placings must be unique")
    state = STATE._load_admin_state()
    override = {"first": payload.first, "second": payload.second, "third": payload.third, "fourth": payload.fourth, "dnf_ids": payload.dnf_ids, "nr_ids": payload.nr_ids, "saved_at": _utc_now().isoformat()}
    state.setdefault("manual_results", {})[race_id] = override
    STATE._save_admin_state(state)
    _audit("admin_results_provisional", {"race_id": race_id})
    return {"ok": True, "race_id": race_id, "saved": True}


@app.post("/api/admin/race/{race_id}/results/confirm_official")
def admin_confirm_results(race_id: str, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    allow_override = os.getenv("FF_ADMIN_SUPER_OVERRIDE", "0") == "1"
    state = STATE._load_admin_state()
    so = state.setdefault("settlement_overrides", {})
    row = so.setdefault(race_id, {"settled_at": None, "official_confirmed": False, "rescore_count": 0})
    if row.get("official_confirmed") and not allow_override:
        raise HTTPException(status_code=409, detail="Official results already confirmed")
    row["official_confirmed"] = True
    row["settled_at"] = _utc_now().isoformat()
    row["rescore_count"] = (safe_int(row.get("rescore_count", 0)) or 0) + 1
    STATE._save_admin_state(state)
    _audit("admin_results_confirm", {"race_id": race_id, "rescore_count": row["rescore_count"]})
    return {"ok": True, "race_id": race_id, "official_confirmed": True, "rescore_count": row["rescore_count"]}


@app.post("/api/admin/upload/racecard_xlsx")
async def admin_upload_racecard(file: UploadFile = File(...), _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    if not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")
    raw = await file.read()
    tmp = DATA / "_upload_tmp.xlsx"
    tmp.write_bytes(raw)
    from openpyxl import load_workbook

    wb = load_workbook(filename=str(tmp), data_only=True)
    ws = wb.active
    races: list[dict[str, Any]] = []
    current: dict[str, Any] | None = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[0]).strip() if row and row[0] is not None else ""
        if not name:
            continue
        header = parse_race_header(name)
        if header:
            idx, off = header
            current = {"race_id": f"manual_{idx}", "name": f"Race {idx}", "off_time": off, "runners": []}
            races.append(current)
            continue
        if current is None:
            continue
        odd_raw = row[1] if len(row) > 1 else None
        dec = parse_odds_to_decimal(odd_raw)
        current["runners"].append({"horse_name": name, "fractional": str(odd_raw) if odd_raw is not None else "", "decimal": dec or 0})
    if not races:
        raise HTTPException(status_code=400, detail="No races found. Ensure headers like 'Race 1- 13:50' and horse rows exist.")
    preview = {"meeting_date": _utc_now().astimezone(STATE.local_tz).date().isoformat(), "course": STATE.target_course, "races": races}
    _atomic_write_json(UPLOAD_PREVIEW_PATH, preview)
    _audit("admin_upload_preview", {"file": file.filename, "races": len(races)})
    return preview


@app.post("/api/admin/upload/commit")
def admin_upload_commit(_: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    preview = _read_json(UPLOAD_PREVIEW_PATH, None)
    if not preview:
        raise HTTPException(status_code=400, detail="No upload preview found")
    state = STATE._load_admin_state()
    state["manual_upload"] = preview
    STATE._save_admin_state(state)
    _audit("admin_upload_commit", {"races": len(preview.get("races", []))})
    return {"ok": True, "source": "manual_upload", "races": len(preview.get("races", []))}


@app.post("/api/admin/invites/create")
def admin_invites_create(payload: InviteCreateIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    if payload.count < 1 or payload.count > 200:
        raise HTTPException(status_code=400, detail="count must be between 1 and 200")
    now = _utc_now()
    invites = _read_json(INVITES_PATH, [])
    new = []
    for _i in range(payload.count):
        token = secrets.token_urlsafe(16)
        row = {"token": token, "created_at": now.isoformat(), "expires_at": (now + timedelta(hours=payload.expires_hours)).isoformat(), "used_at": None, "redeemed_user_id": None, "revoked_at": None}
        invites.append(row)
        new.append(row)
    _atomic_write_json(INVITES_PATH, invites)
    _audit("admin_invites_create", {"count": payload.count, "expires_hours": payload.expires_hours})
    return {"ok": True, "tokens": new}


@app.get("/api/admin/invites/list")
def admin_invites_list(_: dict[str, Any] = Depends(require_admin)) -> list[dict[str, Any]]:
    return _read_json(INVITES_PATH, [])


@app.post("/api/admin/invites/revoke")
def admin_invites_revoke(payload: InviteRevokeIn, _: dict[str, Any] = Depends(require_admin)) -> dict[str, Any]:
    invites = _read_json(INVITES_PATH, [])
    for row in invites:
        if row.get("token") == payload.token:
            now_iso = _utc_now().isoformat()
            row["revoked_at"] = now_iso
            row["used_at"] = row.get("used_at") or now_iso
            _atomic_write_json(INVITES_PATH, invites)
            _audit("admin_invites_revoke", {"token": payload.token})
            return {"ok": True}
    raise HTTPException(status_code=404, detail="Token not found")


@app.post("/api/join")
def join(payload: JoinIn, response: Response) -> dict[str, Any]:
    invites = _read_json(INVITES_PATH, [])
    now = _utc_now()
    target = next((x for x in invites if x.get("token") == payload.token), None)
    if not target:
        raise HTTPException(status_code=404, detail="Invalid token")
    if target.get("revoked_at"):
        raise HTTPException(status_code=409, detail="Token revoked")

    redeemed_user_id = str(target.get("redeemed_user_id") or "").strip() or None
    if redeemed_user_id:
        user = STATE.get_user_by_id(redeemed_user_id)
        if not user:
            raise HTTPException(status_code=409, detail="Token points to a missing user. Ask admin for a fresh invite.")
    else:
        expires_at = target.get("expires_at")
        try:
            if datetime.fromisoformat(str(expires_at)) < now:
                raise HTTPException(status_code=409, detail="Token expired")
        except ValueError:
            raise HTTPException(status_code=409, detail="Token metadata invalid")

        # First redemption creates (or finds) the player account once and binds
        # this invite to that player for later re-entry key behavior.
        user = STATE.add_user(UserIn(display_name=payload.display_name))
        target["redeemed_user_id"] = user["id"]
        target["used_at"] = target.get("used_at") or now.isoformat()
        _atomic_write_json(INVITES_PATH, invites)

    response.set_cookie(
        USER_COOKIE,
        user["id"],
        httponly=True,
        samesite="lax",
        max_age=FF_USER_SESSION_MAX_AGE_SECONDS,
        path="/",
        secure=(os.getenv("FF_COOKIE_SECURE", "0") == "1"),
    )
    _audit("join", {"token": payload.token, "user_id": user["id"]})
    return {"ok": True, "user_id": user["id"], "display_name": user["displayName"], "reentry": bool(redeemed_user_id)}


@app.get("/api/me")
def get_me(user: dict[str, Any] = Depends(require_user)) -> dict[str, Any]:
    return user


@app.get("/api/meeting")
def get_meeting() -> dict[str, Any]:
    snap = STATE.snapshot()
    meeting = snap.get("meeting")
    if not meeting:
        configured_days = STATE.race_day_states or [
            {
                "slot": safe_int(d.get("slot", i + 1)) or (i + 1),
                "course": d.get("course", ""),
                "date": d.get("date", ""),
                "label": d.get("label", ""),
                "status": "pending",
                "races": [],
                "last_refresh": snap.get("last_refresh"),
                "last_error": snap.get("last_error"),
                "next_check_utc": (_utc_now() + timedelta(seconds=STATE.refresh_seconds)).isoformat(),
            }
            for i, d in enumerate(STATE.configured_race_days)
        ]
        configured_days = [_annotate_next_check(d, STATE.refresh_seconds) for d in configured_days]
        return {
            "id": "meeting",
            "course": STATE.target_course,
            "festival": "Fantasy Furlong",
            "days": _meeting_days_from_racedays(configured_days),
            "raceDays": configured_days,
            "snapshotLocked": False,
            "refreshIntervalSeconds": self.refresh_seconds,
        }
    meeting.setdefault("raceDays", STATE.race_day_states or STATE.configured_race_days)
    meeting["raceDays"] = [_annotate_next_check(d, STATE.refresh_seconds) for d in meeting.get("raceDays", [])]
    meeting.setdefault("days", _meeting_days_from_racedays(meeting.get("raceDays", [])))
    meeting.setdefault("refreshIntervalSeconds", STATE.refresh_seconds)
    return meeting




@app.get("/api/configured-racedays")
def get_configured_racedays() -> list[dict[str, Any]]:
    rows = STATE.race_day_states or [
        {
            "slot": safe_int(d.get("slot", i + 1)) or (i + 1),
            "course": d.get("course", ""),
            "date": d.get("date", ""),
            "label": d.get("label", ""),
            "status": "pending",
            "races": [],
            "last_refresh": None,
            "last_error": None,
            "next_check_utc": (_utc_now() + timedelta(seconds=STATE.refresh_seconds)).isoformat(),
        }
        for i, d in enumerate(STATE.configured_race_days)
    ]
    return [_annotate_next_check(d, STATE.refresh_seconds) for d in rows]

@app.get("/api/races")
def get_races() -> list[dict[str, Any]]:
    cached = STATE._route_cache_get("/api/races")
    if cached is not None:
        return cached
    snap = STATE.snapshot()
    races = snap.get("races", [])
    if races:
        STATE._route_cache_set("/api/races", races)
        return races
    state = STATE._load_admin_state()
    manual = state.get("manual_upload")
    if manual:
        meeting, fr = STATE._to_frontend_model({"meeting": {"races": []}})
        _ = meeting
        STATE._route_cache_set("/api/races", fr)
        return fr
    STATE._route_cache_set("/api/races", [])
    return []


@app.get("/api/users")
def get_users() -> list[dict[str, Any]]:
    return STATE.users()


@app.get("/api/picks")
def get_picks() -> list[dict[str, Any]]:
    return STATE.picks()


@app.post("/api/picks")
def post_picks(payload: PickIn, user: dict[str, Any] = Depends(require_user)) -> dict[str, Any]:
    if str(payload.user_id) != str(user.get("id")):
        raise HTTPException(status_code=403, detail="Cannot submit picks for another user")
    STATE.set_pick(payload)
    STATE._route_cache_invalidate()
    return {"ok": True}


@app.post("/api/users")
def post_users(payload: UserIn) -> dict[str, Any]:
    out = STATE.add_user(payload)
    STATE._route_cache_invalidate()
    return out


@app.get("/api/leaderboard")
def get_leaderboard() -> list[dict[str, Any]]:
    cached = STATE._route_cache_get("/api/leaderboard")
    if cached is not None:
        return cached
    out = STATE.leaderboard()
    STATE._route_cache_set("/api/leaderboard", out)
    return out


@app.get("/")
def spa_root():
    return _serve_spa_path("")


@app.get("/{full_path:path}")
def spa_fallback(full_path: str):
    if full_path.startswith("api/") or full_path == "health":
        raise HTTPException(status_code=404, detail="Not Found")
    return _serve_spa_path(full_path)
